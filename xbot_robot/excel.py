# 使用提醒:
# 1. xbot包提供软件自动化、数据表格、Excel、日志、AI等功能
# 2. package包提供访问当前应用数据的功能，如获取元素、访问全局变量、获取资源文件等功能
# 3. 当此模块作为流程独立运行时执行main函数
# 4. 可视化流程中可以通过"调用模块"的指令使用此模块

import xbot
from xbot import print, sleep
from .import package
from .package import variables as glv

def main(args):
    pass
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def process_excel_orders(file_path):
    """
    处理Excel订单数据的函数
   
    参数:
    file_path: Excel文件的路径
   
    功能:
    1. 将G列中的"UK/GB"替换为"GB"
    2. 根据订单号(B列)、客户英文标题(D列)、客户产品变体(E列)合并重复行
    3. 合并时将数量(F列)相加，其他列保持第一次出现时的值
    4. 保存修改后的Excel文件
    5. 将B列和V列设置为文本格式，并根据内容自动调整列宽
    """
   
    # 读取Excel文件，假设第一行是表头
    df = pd.read_excel(file_path, dtype={1: str, 21: str})  # 将B列(索引1)和V列(索引21)设置为字符串类型
   
    # 将G列中的"UK/GB"替换为"GB"
    df.iloc[:, 6] = df.iloc[:, 6].replace('UK/GB', 'GB')  # G列是索引6
   
    # 获取列名，用于后续操作
    columns = df.columns.tolist()
   
    # 定义用于分组的列（B、D、E列，索引分别为1、3、4）
    group_columns = [columns[1], columns[3], columns[4]]  # 订单号、客户英文标题、客户产品变体
    quantity_column = columns[5]  # F列（数量列，索引为5）
   
    # 创建聚合字典
    # 数量列进行求和，其他列取第一个值
    agg_dict = {}
    for col in df.columns:
        if col == quantity_column:
            agg_dict[col] = 'sum'  # 数量列求和
        else:
            agg_dict[col] = 'first'  # 其他列取第一个值
   
    # 按订单号、客户英文标题、客户产品变体分组并聚合
    df_merged = df.groupby(group_columns, as_index=False).agg(agg_dict)
   
    # 保持原始列的顺序
    df_merged = df_merged[columns]
   
    # 保存处理后的Excel文件
    df_merged.to_excel(file_path, index=False)
    
    # 使用openpyxl设置列格式和列宽
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 设置B列和V列为文本格式
    b_col_letter = get_column_letter(2)  # B列
    v_col_letter = get_column_letter(22)  # V列
    
    # 遍历所有行，设置B列和V列的单元格格式为文本
    for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
        b_cell = ws[f'{b_col_letter}{row}']
        v_cell = ws[f'{v_col_letter}{row}']
        
        # 确保值为字符串
        if b_cell.value is not None:
            b_cell.value = str(b_cell.value)
        if v_cell.value is not None:
            v_cell.value = str(v_cell.value)
        
        # 设置单元格数字格式为文本（@符号表示文本格式）
        b_cell.number_format = '@'
        v_cell.number_format = '@'
    
    # 自动调整所有列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        
        for cell in col:
            if cell.value:
                # 计算单元格内容的长度
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # 设置列宽（略微增加宽度以便更好地显示）
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 保存更改
    wb.save(file_path)
   
    print(f"Excel文件处理完成！")
    print(f"原始行数: {len(df)}")
    print(f"合并后行数: {len(df_merged)}")
    print(f"文件已保存到: {file_path}")
    print(f"B列和V列已设置为文本格式，所有列宽已自动调整")
   
    return df_merged

# 使用示例：
# 调用函数处理Excel文件
# result = process_excel_orders("你的文件路径.xlsx")