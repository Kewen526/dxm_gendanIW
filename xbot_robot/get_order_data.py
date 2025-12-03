# 使用提醒:
# 1. xbot包提供软件自动化、数据表格、Excel、日志、AI等功能
# 2. package包提供访问当前应用数据的功能，如获取元素、访问全局变量、获取资源文件等功能
# 3. 当此模块作为流程独立运行时执行main函数
# 4. 可视化流程中可以通过"调用模块"的指令使用此模块

import xbot
from xbot import print, sleep
from .import package
from .package import variables as glv

def main(args):
    pass
import pymysql
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_order_data(order_id, sku_list=None, excel_path=None):
    """
    获取订单数据并保存为Excel
    
    参数:
    order_id: 订单ID
    sku_list: 商品SKU列表，长度应与数据行数相同
    excel_path: 保存Excel的路径
    """
    # 数据库连接参数
    db_config = {
        'host': 'rm-j6ce98dcz1z47ee42so.mysql.rds.aliyuncs.com',
        'user': 'gocrm',
        'password': '4ijmvv7U',
        'database': 'gocrm',
        'port': 3306,
        'charset': 'utf8mb4'
    }
    
    try:
        # 连接数据库
        connection = pymysql.connect(**db_config)
        cursor = connection.cursor(pymysql.cursors.DictCursor)  # 使用字典游标
        
        # SQL查询 - 包含所有需要的字段，新增shop_num字段
        sql = """
        SELECT 
            waybill_num,
            order_id,
            package_id,
            customer_english_title,
            stop_product_variant,
            num,
            nation,
            ali_link,
            channel,
            shipments_remark,
            ali_buyer_word,
            purchase_remark,
            special_pairing,
            remark,
            huo_remark,
            info,
            shop_num
        FROM 
            `orders`
        WHERE 
            order_id = %s
        """
        
        # 执行查询
        cursor.execute(sql, (order_id,))
        
        # 获取结果
        row = cursor.fetchone()
        
        if not row:
            return [["没有找到对应订单号的数据"]]
        
        # 这些字段是列表类型
        array_fields = [
            'customer_english_title', 
            'stop_product_variant', 
            'num', 
            'shipments_remark', 
            'ali_buyer_word', 
            'purchase_remark', 
            'special_pairing', 
            'remark',
            'ali_link'
        ]
        
        # 处理列表字段
        for field in array_fields:
            if field in row and isinstance(row[field], str):
                try:
                    row[field] = json.loads(row[field])
                except json.JSONDecodeError:
                    # 如果不是有效的JSON，保持为字符串但放入列表
                    row[field] = [row[field]]
            elif field in row and not isinstance(row[field], list):
                # 如果不是列表，则放入列表
                row[field] = [row[field]]
        
        # 解析info字段
        if 'info' in row and row['info']:
            if isinstance(row['info'], str):
                try:
                    info_data = json.loads(row['info'])
                except json.JSONDecodeError:
                    info_data = {}
            else:
                info_data = row['info']
        else:
            info_data = {}
        
        # 确定数组的最大长度
        max_length = 0
        for field in array_fields:
            if field in row and isinstance(row[field], list):
                max_length = max(max_length, len(row[field]))
        
        if max_length == 0:
            return [["没有找到数组数据"]]
        
        # ===== 修改后的店铺判断规则 =====
        order_id_value = str(row.get('order_id', ''))
        if 'QXL-' in order_id_value:
            shop_name = "SG-QXL-B专属店铺"
        elif 'SP00001' in order_id_value:
            shop_name = "SP00001专属店铺"
        elif 'SP00002' in order_id_value:
            shop_name = "SP00002专属店铺"
        else:
            shop_name = "手工订单"
        # ================================
        
        # 根据映射关系定义列顺序和对应的中文表头
        columns_mapping = [
            ("运单号", "waybill_num"),
            ("订单号", "order_id"),
            ("包裹号", "package_id"),
            ("客户英文标题", "customer_english_title"),
            ("客户产品变体", "stop_product_variant"),
            ("数量", "num"),
            ("国家", "nation"),
            ("1688链接", "ali_link"),
            ("物流渠道", "channel"),
            ("特殊备注", "shipments_remark"),
            ("1688买家留言", "ali_buyer_word"),
            ("采购单备注", "purchase_remark"),
            ("特殊配对", "special_pairing"),
            ("产品拣货备注", "remark"),
            ("订单拣货备注", "huo_remark"),
            ("邮箱", "info_Email"),
            ("收件人姓名", "info_Name"),
            ("地址", "info_Addr"),
            ("城市", "info_City"),
            ("邮编", "info_Postcode"),
            ("省州", "info_Province"),
            ("电话", "info_Phone"),
            ("属性", "combined_title"),
            ("商品SKU", "product_sku"),
            ("店铺", "shop"),
            ("单价", "price")
        ]
        
        # 提取中文表头和字段名
        chinese_headers = [item[0] for item in columns_mapping]
        field_names = [item[1] for item in columns_mapping]
        
        # 创建结果二维列表
        result = []
        
        # 创建表头行
        result.append(chinese_headers)
        
        # 准备SKU列表
        if not sku_list:
            sku_list = [""] * max_length
        elif len(sku_list) < max_length:
            sku_list.extend([""] * (max_length - len(sku_list)))
        
        # 创建数据行
        for i in range(max_length):
            data_row = []
            
            # 按照映射关系添加数据
            for field in field_names:
                if field == "waybill_num":
                    data_row.append(row.get('waybill_num', ""))
                elif field == "order_id":
                    data_row.append(str(row.get('order_id', "")))  # 确保订单号为字符串
                elif field == "package_id":
                    data_row.append(row.get('package_id', ""))
                elif field == "nation":
                    data_row.append(row.get('nation', ""))
                elif field == "channel":
                    data_row.append(row.get('channel', ""))
                elif field == "huo_remark":
                    data_row.append(row.get('huo_remark', ""))
                elif field in array_fields:  # 列表字段
                    if field in row and isinstance(row[field], list) and i < len(row[field]):
                        data_row.append(row[field][i])
                    else:
                        data_row.append("")
                elif field == "info_Email":
                    data_row.append(info_data.get("Email", ""))
                elif field == "info_Name":
                    data_row.append(info_data.get("Name", ""))
                elif field == "info_Addr":
                    data_row.append(info_data.get("Addr", ""))
                elif field == "info_City":
                    data_row.append(info_data.get("City", ""))
                elif field == "info_Postcode":
                    # 邮编也转换为字符串，防止前导0丢失
                    postcode = info_data.get("Postcode", "")
                    data_row.append(str(postcode) if postcode else "")
                elif field == "info_Province":
                    data_row.append(info_data.get("Province", ""))
                elif field == "info_Phone":
                    # 电话号码必须转换为字符串，保留前导0
                    phone = info_data.get("Phone", "")
                    data_row.append(str(phone) if phone else "")
                elif field == "combined_title":
                    customer_title = row['customer_english_title'][i] if (
                        'customer_english_title' in row and 
                        isinstance(row['customer_english_title'], list) and 
                        i < len(row['customer_english_title'])
                    ) else ""
                    
                    product_variant = row['stop_product_variant'][i] if (
                        'stop_product_variant' in row and 
                        isinstance(row['stop_product_variant'], list) and 
                        i < len(row['stop_product_variant'])
                    ) else ""
                    
                    combined_title = f"{customer_title}-{product_variant}" if customer_title and product_variant else ""
                    data_row.append(combined_title)
                elif field == "product_sku":
                    # 从传入的SKU列表获取
                    data_row.append(sku_list[i] if i < len(sku_list) else "")
                elif field == "shop":
                    # 根据order_id判断店铺类型
                    data_row.append(shop_name)
                elif field == "price":
                    # 单价固定为【15】，确保为字符串
                    data_row.append("15")
            
            result.append(data_row)
        
        # 如果指定了Excel路径，则保存为Excel文件并进行格式化
        if excel_path:
            df = pd.DataFrame(result[1:], columns=result[0])  # 第一行作为列名
            
            # 将需要保持为文本格式的列转换为字符串，防止科学计数法显示
            df['订单号'] = df['订单号'].astype(str)
            df['电话'] = df['电话'].astype(str)  # 确保电话为字符串格式
            df['邮编'] = df['邮编'].astype(str)  # 邮编也可能有前导0
            df['单价'] = df['单价'].astype(str)
            
            # 使用ExcelWriter保存，以便后续格式化
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # 设置列宽自适应
                for idx, col in enumerate(df.columns):
                    column_width = max(
                        df[col].astype(str).map(len).max(),  # 最长数据的长度
                        len(col) + 2  # 表头长度加上一些额外空间
                    )
                    # 设置列宽，确保至少有一个最小宽度
                    column_width = max(column_width, 10)
                    # 转换为Excel列标识（A, B, C...）
                    column_letter = get_column_letter(idx + 1)
                    # 设置列宽
                    worksheet.column_dimensions[column_letter].width = column_width
                
                # 设置特定列为文本格式，防止数字被转换
                text_format_columns = {
                    2: "订单号",      # B列 - 订单号
                    20: "邮编",       # T列 - 邮编  
                    22: "电话",       # V列 - 电话
                    26: "单价"        # Z列 - 单价
                }
                
                for row in range(2, len(df) + 2):  # Excel行从1开始，且有表头
                    for col_num, col_name in text_format_columns.items():
                        cell = worksheet.cell(row=row, column=col_num)
                        cell.number_format = '@'  # 设置为文本格式
                        
                        # 特别处理电话列，确保值被正确写入
                        if col_name == "电话":
                            phone_value = str(df.iloc[row-2]['电话'])  # row-2因为DataFrame索引从0开始
                            cell.value = phone_value
            
            print(f"数据已保存到: {excel_path}")
        
        return result
    
    except Exception as e:
        error_result = [["错误", str(e)]]
        print(f"发生错误: {str(e)}")
        return error_result
    
    finally:
        if 'connection' in locals() and connection:
            cursor.close()
            connection.close()

# 使用示例
if __name__ == "__main__":
    order_id = input("请输入订单号: ")
    
    # 是否需要指定SKU
    use_sku = input("是否需要指定SKU列表? (y/n): ").lower()
    sku_list = None
    if use_sku == 'y':
        sku_input = input("请输入SKU列表，用逗号分隔: ")
        sku_list = [sku.strip() for sku in sku_input.split(",")]
    
    # 是否需要保存为Excel
    save_excel = input("是否需要保存为Excel? (y/n): ").lower()
    excel_path = None
    if save_excel == 'y':
        excel_path = input("请输入Excel保存路径: ")
    
    # 获取数据
    result = get_order_data(order_id, sku_list, excel_path)
    
    # 打印结果
    print("\n查询结果:")
    for i, row in enumerate(result):
        if i == 0:
            print("表头:", row)
        else:
            print(f"行 {i}:", row)