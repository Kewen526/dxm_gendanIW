# 使用提醒:
# 1. xbot包提供软件自动化、数据表格、Excel、日志、AI等功能
# 2. package包提供访问当前应用数据的功能，如获取元素、访问全局变量、获取资源文件等功能
# 3. 当此模块作为流程独立运行时执行main函数
# 4. 可视化流程中可以通过"调用模块"的指令使用此模块

import xbot
from xbot import print, sleep
from .import package
from .package import variables as glv

def main(args):
    pass

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def process_excel_orders_detailed(file_path):
    """
    处理Excel订单数据的函数（详细显示处理过程）
   
    参数:
    file_path: Excel文件的路径
   
    功能:
    1. 将G列中的"UK/GB"替换为"GB"
    2. 根据订单号(B列)、客户英文标题(D列)、客户产品变体(E列)合并重复行
    3. 合并时将数量(F列)相加，保留第一次出现行的位置，删除后续重复行
    4. 详细显示哪些行被保留，哪些行被删除
    5. 处理电话号码列(V列)，保留前导0，防止科学计数法显示
    """
   
    # 先读取Excel获取列名结构
    temp_df = pd.read_excel(file_path, nrows=1)
    columns = temp_df.columns.tolist()
    
    # 设置需要保持为字符串格式的列名（防止前导0丢失）
    dtype_dict = {}
    
    # B列 - 订单号 (索引1)
    if len(columns) > 1:
        dtype_dict[columns[1]] = str
    
    # T列 - 邮编 (索引19) 
    if len(columns) > 19:
        dtype_dict[columns[19]] = str
        
    # V列 - 电话 (索引21)
    if len(columns) > 21:
        dtype_dict[columns[21]] = str
        
    # Z列 - 单价 (索引25，如果存在)
    if len(columns) > 25:
        dtype_dict[columns[25]] = str
    
    print(f"设置为字符串类型的列: {list(dtype_dict.keys())}")
    
    # 使用列名指定数据类型来读取Excel文件，保持前导0
    df = pd.read_excel(file_path, dtype=dtype_dict)
   
    print(f"读取Excel文件: {file_path}")
    print(f"原始数据行数: {len(df)} 行")
    print("-" * 50)
   
    # 检查和处理NaN值
    if len(df.columns) > 21:  # 检查V列（电话）
        phone_column = df.columns[21]
        df[phone_column] = df[phone_column].fillna('')  # 将NaN替换为空字符串
        print(f"电话列({phone_column})已处理NaN值，保持字符串格式")
   
    if len(df.columns) > 19:  # 检查T列（邮编）
        postcode_column = df.columns[19]
        df[postcode_column] = df[postcode_column].fillna('')  # 将NaN替换为空字符串
        print(f"邮编列({postcode_column})已处理NaN值，保持字符串格式")
   
    # 将G列中的"UK/GB"替换为"GB"
    uk_gb_count = (df.iloc[:, 6] == 'UK/GB').sum()
    df.iloc[:, 6] = df.iloc[:, 6].replace('UK/GB', 'GB')
    if uk_gb_count > 0:
        print(f"已将 {uk_gb_count} 个 'UK/GB' 替换为 'GB'")
   
    # 获取列名
    columns = df.columns.tolist()
    group_columns = [columns[1], columns[3], columns[4]]  # B、D、E列
    quantity_column = columns[5]  # F列
   
    print(f"分组依据列: {group_columns}")
    print(f"数量汇总列: {quantity_column}")
    print("-" * 50)
   
    # 创建处理结果的副本
    df_result = df.copy()
   
    # 创建分组键来识别重复行
    df_result['_group_key'] = df_result[group_columns].astype(str).agg('_'.join, axis=1)
   
    # 统计每个组的出现次数
    group_counts = df_result['_group_key'].value_counts()
    duplicate_groups = group_counts[group_counts > 1]
   
    print(f"发现 {len(duplicate_groups)} 个重复组:")
    
    total_deleted = 0
    merge_details = []
   
    for group_key, count in duplicate_groups.items():
        # 获取当前组的所有行索引
        group_mask = df_result['_group_key'] == group_key
        group_indices = df_result[group_mask].index.tolist()
        
        # 获取第一行的信息用于显示
        first_row = df_result.loc[group_indices[0]]
        group_info = f"{first_row[group_columns[0]]}-{first_row[group_columns[1]]}-{first_row[group_columns[2]]}"
        
        # 计算总数量
        quantities = df_result.loc[group_indices, quantity_column].tolist()
        total_quantity = sum(quantities)
        
        # 保留第一行，删除其他行
        keep_row = group_indices[0] + 2  # +2因为Excel从第2行开始（第1行是表头）
        delete_rows = [i + 2 for i in group_indices[1:]]
        
        # 更新第一行的数量
        df_result.loc[group_indices[0], quantity_column] = total_quantity
        
        print(f"  组 {len(merge_details)+1}: {group_info[:60]}...")
        print(f"    保留: 第{keep_row}行 (数量: {quantities[0]} → {total_quantity})")
        print(f"    删除: 第{delete_rows}行 (数量: {quantities[1:]})")
        
        merge_details.append({
            'group': group_info,
            'keep_row': keep_row,
            'delete_rows': delete_rows,
            'original_quantities': quantities,
            'final_quantity': total_quantity
        })
        
        total_deleted += len(group_indices) - 1
   
    # 删除重复行
    duplicate_indices = []
    for group_key in duplicate_groups.index:
        group_mask = df_result['_group_key'] == group_key
        group_indices = df_result[group_mask].index.tolist()
        duplicate_indices.extend(group_indices[1:])  # 保留第一个，删除其他
   
    df_result = df_result.drop(index=duplicate_indices)
    
    # 移除辅助列
    df_result = df_result.drop('_group_key', axis=1)
    
    # 重置索引
    df_result = df_result.reset_index(drop=True)
   
    print("-" * 50)
    print(f"处理结果:")
    print(f"  原始行数: {len(df)}")
    print(f"  最终行数: {len(df_result)}")
    print(f"  删除行数: {total_deleted}")
    print(f"  保留行数: {len(df) - total_deleted}")
   
    # 保存文件，使用openpyxl引擎以便格式化
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_result.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # 设置需要保持为文本格式的列，防止科学计数法显示
        text_format_columns = {}
        
        # B列 - 订单号
        if len(df_result.columns) > 1:
            text_format_columns[2] = "订单号"
            
        # T列 - 邮编 (索引19)
        if len(df_result.columns) > 19:
            text_format_columns[20] = "邮编"
            
        # V列 - 电话 (索引21) 
        if len(df_result.columns) > 21:
            text_format_columns[22] = "电话"
            
        # Z列 - 单价 (如果存在)
        if len(df_result.columns) > 25:
            text_format_columns[26] = "单价"
        
        print(f"设置文本格式的列: {text_format_columns}")
        
        # 设置指定列为文本格式
        for row in range(2, len(df_result) + 2):  # Excel行从1开始，且有表头
            for col_num, col_name in text_format_columns.items():
                if col_num <= len(df_result.columns):  # 确保列存在
                    cell = worksheet.cell(row=row, column=col_num)
                    cell.number_format = '@'  # 设置为文本格式
                    
                    # 获取并设置单元格值，确保为字符串格式
                    if col_name == "电话" and len(df_result.columns) > 21:
                        phone_value = df_result.iloc[row-2, 21]  # 第22列，索引21
                        # 确保电话号码为字符串，并处理空值
                        cell.value = str(phone_value) if pd.notna(phone_value) and phone_value != '' else ''
                    elif col_name == "邮编" and len(df_result.columns) > 19:
                        postcode_value = df_result.iloc[row-2, 19]  # 第20列，索引19
                        # 确保邮编为字符串，并处理空值
                        cell.value = str(postcode_value) if pd.notna(postcode_value) and postcode_value != '' else ''
                    elif col_name == "订单号":
                        order_value = df_result.iloc[row-2, 1]  # 第2列，索引1
                        # 确保订单号为字符串，并处理空值
                        cell.value = str(order_value) if pd.notna(order_value) and order_value != '' else ''
                    elif col_name == "单价" and len(df_result.columns) > 25:
                        price_value = df_result.iloc[row-2, 25]  # 第26列，索引25
                        # 确保单价为字符串，并处理空值
                        cell.value = str(price_value) if pd.notna(price_value) and price_value != '' else ''
        
        # 设置列宽自适应
        for idx, col in enumerate(df_result.columns):
            column_width = max(
                df_result[col].astype(str).map(len).max(),  # 最长数据的长度
                len(col) + 2  # 表头长度加上一些额外空间
            )
            # 设置列宽，确保至少有一个最小宽度
            column_width = max(column_width, 10)
            # 转换为Excel列标识（A, B, C...）
            column_letter = get_column_letter(idx + 1)
            # 设置列宽
            worksheet.column_dimensions[column_letter].width = column_width
    
    print(f"文件已保存到: {file_path}")
    print("已设置电话号码和邮编列为文本格式，保留前导0")
   
    return df_result, merge_details

# 使用示例：
# result, details = process_excel_orders_detailed("你的文件路径.xlsx")
# 
# # 如果需要查看合并详情
# for detail in details:
#     print(f"组: {detail['group']}")
#     print(f"保留第{detail['keep_row']}行，删除第{detail['delete_rows']}行")